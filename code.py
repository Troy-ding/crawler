import requests as re
from bs4 import BeautifulSoup, element
import pandas as pd
from joblib import Parallel, delayed
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def parse_card(card: element.Tag):
    """
    解析网页中的单张卡片
    :param card: 卡片标签
    :return:
    """
    card_title = card.find_all(name='h4', attrs='showcaseCardTitle_zvaY')[0].find('a').text
    card_body = card.find_all(name='p', attrs='showcaseCardBody_fqoj')
    card_tips = card_body[0].text
    card_prompt = card_body[1].text
    
    card_foot_rs = card.find_all(name='li', attrs={'class': 'tag_dHH4'})
    card_foot = '|'.join([cfr.text for cfr in card_foot_rs])
    return card_title, card_foot, card_tips, card_prompt
    
def parse_url(url: str):
    """
    解析url，返回网页信息
    :param url: 网页链接
    :return:
    """
    response = re.get(url)

    status_code = response.status_code
    if status_code != 200:
        return None
    
    bs = BeautifulSoup(response.content, 'lxml')
    cards = bs.find_all(name='li', attrs={'class': 'card shadow--md'})
    
    df = pd.DataFrame(data=[parse_card(c) for c in cards],
                      columns=['title', 'class', 'tips', 'prompt'])
    return df

def save_excel(df: pd.DataFrame, output_path: str, index_label=None):
    """
    保存网页信息
    :param df: 待保存表格
    :param output_path: 表格路径
    :return:
    """
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, sheet_name='Sheet1', index=True, index_label=index_label)
    worksheet = writer.sheets['Sheet1']

    index_size = len(df.index[0])
    # 设置单元格宽度
    for i, col in enumerate(df.columns):
        max_len = df[col].astype(str).map(len).max()
        col_width = min(max((max_len // 15), 1) * 15, 60)
        worksheet.column_dimensions[worksheet.cell(row=1, column=i + 1 + index_size).column_letter].width = col_width

    # 设置自动换行
    for row in worksheet.iter_rows():
        for cell in row[index_size:]:
            cell.alignment = cell.alignment.copy(wrap_text=True, horizontal='left', vertical='center')
            
    for row in worksheet.iter_rows():
        for cell in row[:index_size]:
            cell.alignment = cell.alignment.copy(wrap_text=True, horizontal='center', vertical='center')
    writer.save()
    writer.close()

url = 'https://prompt-shortcut.writeathon.cn/'
url_cn = 'https://prompt-shortcut.writeathon.cn/cn'

df_en = parse_url(url)
df_cn = parse_url(url_cn)

df = df_en.copy()
df['prompt_cn'] = df_cn['prompt']

# 提示器分类
category = df['class'].apply(lambda x: x.split('|'))
category = set([element for sublist in category.tolist() for element in sublist])
df_category = [df[df['class'].str.contains(cg)].reset_index(drop=True) for cg in category]
result = pd.concat(dict(zip(category, df_category)), axis=0, keys=category).drop(columns=['class'])

# 保存结果
save_excel(result, 'output.xlsx', ['class', 'No'])
